"""
Tech Events Scraper — v2
=========================
Finds upcoming tech conferences, summits, hackathons, and expos across four
community lenses: Black Tech, Women in Tech, LGBTQ+ Tech, and General Tech.

Key features
------------
* Brave Search API — whole-web search, no CAPTCHA
* Community tagging — each event labelled Black Tech / Women in Tech /
  LGBTQ+ Tech / General (multiple tags possible)
* Past-event filter — drops events whose date has already passed
* Scope: conferences/summits/hackathons only (not weekly meetups or webinars)
* Sources: Brave API, Eventbrite, Meetup, Luma
* Exports to a timestamped local Excel file (always) and optionally Google Sheets
* Fully non-interactive — no prompts during a run.

Quick start
-----------
1.  Copy .env.example → .env and fill in BRAVE_API_KEY
    (see .env.example for step-by-step instructions)
2.  pip install -r requirements.txt && playwright install chromium
3.  python scraper.py

Output
------
* black_tech_events_YYYYMMDD_HHMMSS.xlsx  (always written)
* Google Sheet link printed to console    (if ENABLE_GOOGLE_SHEETS=true)
"""

import asyncio
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

load_dotenv()

# ── Configuration ─────────────────────────────────────────────────────────────

BRAVE_API_KEY = os.getenv("BRAVE_API_KEY", "")
ENABLE_SHEETS = os.getenv("ENABLE_GOOGLE_SHEETS", "false").lower() == "true"

# Auth method for Google Sheets: "oauth" or "service_account"
SHEETS_AUTH_METHOD   = "oauth"
CREDENTIALS_FILE     = "credentials.json"
SERVICE_ACCOUNT_FILE = "service_account.json"
TOKEN_FILE           = "token.json"

SHEET_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ── Event store ───────────────────────────────────────────────────────────────
EVENT_STORE_PATH = Path("data/events.json")
EVENT_EXPIRY_DAYS = 90   # drop events not seen for this many days
NEW_BADGE_DAYS   = 7     # show "NEW" badge if first_seen within this window

# ── Search queries ────────────────────────────────────────────────────────────
# Focused on conference/summit/hackathon scale events — AfroTech / ReactATL tier.
# Year anchored to current + next so we don't pull stale results.

_YEAR = datetime.now().year
_YEARS = f"{_YEAR} OR {_YEAR + 1}"


# Tier 1 — site-targeted Brave queries (highest precision, direct event pages)
BRAVE_SITE_QUERIES = [
    # Black Tech
    f"black tech conference {_YEARS} site:eventbrite.com",
    f"black developer summit {_YEARS} site:eventbrite.com",
    f"black in tech hackathon {_YEARS} site:eventbrite.com",
    f"hbcu tech {_YEARS} site:eventbrite.com",
    f"black founders {_YEARS} site:eventbrite.com",
    f"colorstack {_YEARS} site:eventbrite.com",
    f"nsbe {_YEARS} site:eventbrite.com",
    f"black tech {_YEARS} site:lu.ma",
    f"afrotech {_YEARS} site:lu.ma",
    f"colorstack {_YEARS} site:lu.ma",
    f"code2040 {_YEARS} site:lu.ma",
    # Black Tech — Devpost hackathons
    f"black tech hackathon {_YEARS} site:devpost.com",
    f"hbcu hackathon {_YEARS} site:devpost.com",
    # Women in Tech
    f"women in tech conference {_YEARS} site:eventbrite.com",
    f"women developer summit {_YEARS} site:eventbrite.com",
    f"girls in tech {_YEARS} site:eventbrite.com",
    f"lesbians who tech {_YEARS} site:eventbrite.com",
    f"women in tech {_YEARS} site:lu.ma",
    f"women in tech hackathon {_YEARS} site:devpost.com",
    # LGBTQ+ Tech
    f"lgbtq tech conference {_YEARS} site:eventbrite.com",
    f"queer tech {_YEARS} site:eventbrite.com",
    f"out in tech {_YEARS} site:eventbrite.com",
    f"lgbtq tech {_YEARS} site:lu.ma",
    # General Tech
    f"tech conference {_YEARS} site:eventbrite.com",
    f"developer conference {_YEARS} site:eventbrite.com",
    f"startup conference {_YEARS} site:eventbrite.com",
    f"tech summit {_YEARS} site:lu.ma",
    f"developer hackathon {_YEARS} site:eventbrite.com",
    f"diversity hackathon {_YEARS} site:devpost.com",
]

# Tier 2 — whole-web Brave queries (broader reach, stricter post-filter applied)
BRAVE_QUERIES = [
    # Black Tech — named events & orgs
    f"Black tech conference {_YEARS}",
    f"Black developer conference {_YEARS}",
    f"HBCU tech conference {_YEARS}",
    f"Black founders conference {_YEARS}",
    f"AfroTech {_YEARS}",
    f"BlackTechWeek {_YEARS}",
    f"ColorStack summit {_YEARS}",
    f"NSBE convention {_YEARS}",
    f"Code2040 {_YEARS}",
    f"Black in AI {_YEARS}",
    # Women in Tech
    f"women in tech conference {_YEARS}",
    f"Grace Hopper Celebration {_YEARS}",
    f"women developer summit {_YEARS}",
    f"Lesbians Who Tech {_YEARS}",
    # LGBTQ+ Tech
    f"LGBTQ tech conference {_YEARS}",
    f"Out in Tech conference {_YEARS}",
    # General Tech — anchored to US/Canada to avoid pulling international results
    f"software engineering conference {_YEARS} United States",
    f"developer conference {_YEARS} United States",
    f"startup tech summit {_YEARS} United States",
    f"AI conference {_YEARS} United States",
    f"tech hackathon {_YEARS} United States OR Canada",
    f"MLH hackathon {_YEARS} United States",
    f"tech week {_YEARS} United States",
]

EVENTBRITE_QUERIES = [
    # Black Tech
    "black tech conference",
    "black developer summit",
    "afrotech",
    "hbcu tech",
    "colorstack",
    "nsbe conference",
    # Women in Tech
    "women in tech conference",
    "women developer",
    "lesbians who tech",
    # LGBTQ+
    "lgbtq tech",
    "queer tech",
    # General
    "tech conference",
    "developer summit",
    "startup conference",
]

MEETUP_QUERIES = [
    "black tech conference",
    "women in tech conference",
    "lgbtq tech",
    "developer conference",
]

LUMA_QUERIES = [
    "black tech",
    "afrotech",
    "colorstack",
    "women in tech",
    "lgbtq tech",
    "tech conference",
    "developer summit",
]

DEVPOST_QUERIES = [
    "black tech",
    "diversity",
    "women tech",
    "lgbtq",
    "hbcu",
]

# Topics fetched from the Confs.tech open dataset (GitHub-hosted JSON)
CONFS_TECH_TOPICS = [
    "javascript", "python", "devops", "security", "data", "general", "ux",
]
_CONFS_TECH_NA_COUNTRIES = {
    "U.S.A.", "USA", "United States", "United States of America", "Canada",
}

# ── Community signals ─────────────────────────────────────────────────────────

_BLACK_SIGNALS = {
    "black", "afro", "african american", "hbcu", "melanin", "poc tech", "nsbe",
    "afrotech", "blacktech", "black tech", "colorstack", "code2040", "black in ai",
    "blacktechweek", "black founders",
}
_WOMEN_SIGNALS = {
    "women in tech", "women tech", "women developer", "girls in tech",
    "girl in tech", "grace hopper", "women who code", "lesbians who tech",
    "women in stem", "female founder", "women founder",
}
_LGBTQ_SIGNALS = {
    "lgbtq", "lgbt", "queer tech", "queer in tech", "out in tech",
    "lesbians who tech", "pride tech", "trans in tech",
}
_TECH_TERMS = {
    "tech", "technology", "software", "developer", "coding", "engineering",
    "startup", "ai ", "machine learning", "data science", "cybersecurity",
    "cloud", "devops", "blockchain", "fintech", "saas", "open source",
    "product", "ux", "design", "innovation",
}

# ── North America filter ──────────────────────────────────────────────────────
# Matches unambiguous country names that indicate a non-NA event.
# Deliberately excludes ambiguous city names (London/Dublin/Paris also exist in NA).
_INTL_LOCATION_RE = re.compile(
    r"\b(united kingdom|england|scotland|wales|australia|new zealand|"
    r"germany|france|spain|italy|netherlands|sweden|norway|denmark|"
    r"finland|switzerland|austria|portugal|ireland|poland|india|singapore|"
    r"japan|china|hong kong|south korea|taiwan|nigeria|kenya|south africa|"
    r"united arab emirates|brazil|argentina|colombia)\b",
    re.IGNORECASE,
)

# ccTLDs that reliably mean a non-NA host domain
_INTL_TLD_RE = re.compile(
    r"https?://[^/]*\.(uk|au|nz|de|fr|es|it|nl|be|se|no|dk|fi|ch|at|pt|ie|"
    r"pl|in|sg|jp|cn|hk|kr|tw|ng|ke|za|br|ar)(?:/|$)",
    re.IGNORECASE,
)


def _is_north_america(event: dict) -> bool:
    """
    Return False when location + title or the URL TLD clearly points to a
    non-North-America event.  Permissive — only rejects on strong evidence so
    events with missing/ambiguous location are kept.
    """
    loc_title = f"{event.get('location', '')} {event.get('title', '')}"
    if _INTL_LOCATION_RE.search(loc_title):
        return False
    url = event.get("link", "")
    if url and _INTL_TLD_RE.search(url):
        return False
    return True


# ── Conference type & exclusion terms ────────────────────────────────────────

CONFERENCE_TYPE_TERMS = {
    "conference", "conf", "summit", "hackathon", "expo", "symposium",
    "convention", "gala", "forum", "workshop series", "bootcamp",
    "pitch competition", "tech week",
}

EXCLUDE_CONTENT_TERMS = {
    "weekly meetup", "monthly meetup", "study group", "book club",
    "webinar", "online only", "virtual only",
    "hair", "fashion", "music festival", "food", "comedy",
}

# URL path segments that reliably indicate a news/blog/list page — not an event.
_NON_EVENT_URL_RE = re.compile(
    r"/(blog|news|article|articles|post|posts|category|tag|press|stories|"
    r"insights|resources|guides|roundup|listicle|opinion|features)/",
    re.IGNORECASE,
)

# Domains that publish articles *about* events but are not event platforms.
_ARTICLE_DOMAINS = {
    "medium.com", "forbes.com", "techcrunch.com", "businessinsider.com",
    "theverge.com", "wired.com", "mashable.com", "fastcompany.com",
    "blackenterprise.com", "essence.com", "theroot.com", "blavity.com",
    "afrotech.com/news",  # news subdirectory — keep afrotech.com root
    "linkedin.com", "twitter.com", "x.com", "reddit.com", "quora.com",
    "wikipedia.org", "eventbrite.com/blog",
}

# URL patterns that confirm this IS a direct event page.
_EVENT_URL_RE = re.compile(
    r"("
    r"eventbrite\.com/e/"
    r"|lu\.ma/[^/?#]+"        # lu.ma/slug  (not lu.ma/search or lu.ma/calendar)
    r"|meetup\.com/.+/events/\d+"
    r"|hopin\.com/events/"
    r"|airmeet\.com/e/"
    r"|afrotech\.com(?!/news)"
    r"|reactatl\.com"
    r"|blacktechweek\.com"
    r"|nsbe\.org"
    r"|[a-z0-9][a-z0-9\-]*\.devpost\.com"   # individual hackathon subdomains
    r")",
    re.IGNORECASE,
)

# Title patterns that indicate a roundup/list article rather than an event page.
_LIST_TITLE_RE = re.compile(
    r"(\d+\s+(best|top|must.?attend|black\s+tech|upcoming)\s+"
    r"|^(best|top|must.?attend)\s+"
    r"|events?\s+to\s+(attend|check\s+out|know\s+about|watch)"
    r"|conferences?\s+for\s+"
    r"|guide\s+to\s+"
    r"|roundup"
    r"|everything\s+you\s+need\s+to\s+know)",
    re.IGNORECASE,
)


# ── Community classifier ──────────────────────────────────────────────────────

def classify_community(event: dict) -> str:
    """
    Return a community label for the event based on title + snippet signals.
    Multiple tags are joined with ' / ' (e.g. 'Black Tech / Women in Tech').
    Falls back to 'General' if no community signal is found.
    """
    text = (
        event.get("title", "") + " " + event.get("snippet", "")
    ).lower()

    tags = []
    if any(t in text for t in _BLACK_SIGNALS):
        tags.append("Black Tech")
    if any(t in text for t in _WOMEN_SIGNALS):
        tags.append("Women in Tech")
    if any(t in text for t in _LGBTQ_SIGNALS):
        tags.append("LGBTQ+ Tech")

    return " / ".join(tags) if tags else "General"


# ── Date parsing & upcoming filter ────────────────────────────────────────────

_TODAY = datetime.now().date()

_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

def _parse_event_date(date_str: str) -> date | None:
    """Parse an extracted date string into a date object. Returns None if unparseable."""
    if not date_str:
        return None
    s = date_str.strip()

    # Standard formats with year
    for fmt in (
        "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y",
        "%b. %d, %Y", "%Y-%m-%d", "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # Range like "Nov 15–17, 2026" — use start date
    m = re.match(r"(\w+)\s+(\d{1,2})[–\-]\d{1,2},?\s+(\d{4})", s, re.IGNORECASE)
    if m:
        month_num = _MONTH_ABBREVS.get(m.group(1)[:3].lower())
        if month_num:
            try:
                return date(int(m.group(3)), month_num, int(m.group(2)))
            except ValueError:
                pass

    return None


def is_upcoming(event: dict) -> bool:
    """
    Return True if the event is in the future or its date cannot be determined.
    Only rejects an event when we can confidently parse a past date.
    """
    parsed = _parse_event_date(event.get("date", ""))
    if parsed is None:
        return True          # unknown date — keep it
    return parsed >= _TODAY


def _is_article_domain(url: str) -> bool:
    """Return True if the URL belongs to a known article/news domain."""
    try:
        from urllib.parse import urlparse
        host = urlparse(url).netloc.lower().lstrip("www.")
        return any(host == d or host.endswith("." + d) for d in _ARTICLE_DOMAINS)
    except Exception:
        return False


def is_event_page(url: str, title: str = "") -> bool:
    """
    Return True when the URL or title strongly suggests a direct event page.
    Return False when the URL/title looks like an article, list, or roundup.
    Return None (unknown) for ambiguous cases — caller decides how to treat them.
    """
    if not url:
        return None  # type: ignore[return-value]

    # Confirmed event platform URL → always keep
    if _EVENT_URL_RE.search(url):
        return True

    # Known article domain → always reject
    if _is_article_domain(url) or _NON_EVENT_URL_RE.search(url):
        return False

    # Title looks like a list article → reject
    if title and _LIST_TITLE_RE.search(title):
        return False

    return None  # ambiguous — let is_relevant decide


def is_relevant(event: dict, strict: bool = False) -> bool:
    """
    Return True if the event is a legitimate tech conference/summit/hackathon.

    Community tagging (Black / Women / LGBTQ+ / General) is handled separately
    by classify_community() — this function only gates on event quality.

    strict=True  — general Brave web results (articles mixed in); requires a
                   confirmed event-platform URL plus conference-type signal.
    strict=False — direct scraper results (Eventbrite, Luma, Meetup); trust
                   the source, lighter touch on text signals.
    """
    url   = event.get("link", "")
    title = event.get("title", "")
    text  = " ".join([title, event.get("snippet", ""), event.get("location", "")]).lower()

    # Hard reject: known article/list page
    page_check = is_event_page(url, title)
    if page_check is False:
        return False

    if any(t in text for t in EXCLUDE_CONTENT_TERMS):
        return False

    has_tech = any(t in text for t in _TECH_TERMS)
    has_type = any(t in text for t in CONFERENCE_TYPE_TERMS)

    if strict:
        # General web results: must be a confirmed event URL + tech + conference type
        return has_tech and has_type and (page_check is True)

    # Scraper / site-targeted results: confirmed event URL needs only tech signal;
    # unknown URLs need both tech + conference-type to pass
    if page_check is True:
        return has_tech
    return has_tech and has_type

# ── Helpers ───────────────────────────────────────────────────────────────────

# Ordered date patterns — first match wins.
_DATE_PATTERNS = [
    # "Nov 15–17, 2026" / "November 15, 2026" / "Nov 15 2026"
    re.compile(
        r"\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"\.?\s+\d{1,2}(?:st|nd|rd|th)?"
        r"(?:\s*[–\-]\s*\d{1,2}(?:st|nd|rd|th)?)?"
        r"(?:,?\s+\d{4})?",
        re.IGNORECASE,
    ),
    # ISO: 2026-10-15  or  2026/10/15
    re.compile(r"\b(20\d\d)[.\-/](0[1-9]|1[0-2])[.\-/](0[1-9]|[12]\d|3[01])\b"),
    # US slash: 10/15/2026
    re.compile(r"\b(0?[1-9]|1[0-2])/(0?[1-9]|[12]\d|3[01])/(20\d\d)\b"),
]

# Primary: "City, STATE" where STATE is a US state or Canadian province abbreviation.
_CITY_STATE_RE = re.compile(
    r"\b([A-Z][a-zA-Z .'\-]{1,28}),\s*"
    r"(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|ID|IL|IN|IA|KS|KY|LA|ME|MD|MA|MI|MN|"
    r"MS|MO|MT|NE|NV|NH|NJ|NM|NY|NC|ND|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VT|VA|"
    r"WA|WV|WI|WY|DC|AB|BC|MB|NB|NL|NS|NT|NU|ON|PE|QC|SK|YT)\b"
)

# Fallback: word after "in / at / ·"
_LOC_RE = re.compile(
    r"(?:\·|\bin\b|\bat\b)\s+([A-Z][a-zA-Z .'\-]{2,30}(?:,\s*[A-Z][a-zA-Z ]{1,20})?)"
)


def extract_date(text: str) -> str:
    for pat in _DATE_PATTERNS:
        m = pat.search(text)
        if m:
            return m.group(0).strip()
    return ""


def extract_location(text: str) -> str:
    # Prefer explicit City, State/Province abbreviation (most reliable)
    m = _CITY_STATE_RE.search(text)
    if m:
        return f"{m.group(1).strip()}, {m.group(2)}"
    # Fallback: preposition-based with hard length cap
    m = _LOC_RE.search(text)
    if m:
        loc = m.group(1).strip()
        loc = re.split(r"[.!?]|\s+(?:on|–|\|)\s+", loc)[0].strip()
        return loc[:50]
    return ""


def deduplicate(results: list[dict]) -> list[dict]:
    seen, unique = set(), []
    for r in results:
        # Prefer URL (strip query string) as the canonical dedup key; fall back to title
        key = r.get("link", "").split("?")[0].lower() or r["title"].lower().strip()
        if key and key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


# ── Persistent event store ────────────────────────────────────────────────────

def load_event_store() -> dict:
    if EVENT_STORE_PATH.exists():
        try:
            return json.loads(EVENT_STORE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def merge_into_store(store: dict, events: list[dict], today: str) -> dict:
    for event in events:
        key = event.get("link", "").split("?")[0].lower() or event.get("title", "").lower().strip()
        if not key:
            continue
        if key in store:
            store[key]["last_seen"]   = today
            store[key]["appearances"] = store[key].get("appearances", 1) + 1
            # Overwrite community (classifier may improve over time)
            if event.get("community"):
                store[key]["community"] = event["community"]
            # Fill in missing fields if we have better data now
            for field in ("date", "location", "title"):
                if event.get(field) and not store[key].get(field):
                    store[key][field] = event[field]
        else:
            store[key] = {**event, "first_seen": today, "last_seen": today, "appearances": 1}
    return store


def expire_past_events(store: dict, today: str) -> dict:
    today_date = date.fromisoformat(today)
    active = {}
    for key, event in store.items():
        last_seen    = date.fromisoformat(event.get("last_seen", today))
        days_unseen  = (today_date - last_seen).days
        parsed_date  = _parse_event_date(event.get("date", ""))
        if days_unseen > EVENT_EXPIRY_DAYS:
            continue
        if parsed_date and parsed_date < today_date and days_unseen > 7:
            continue
        active[key] = event
    return active


def save_event_store(store: dict) -> None:
    EVENT_STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    EVENT_STORE_PATH.write_text(
        json.dumps(store, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def store_to_list(store: dict) -> list[dict]:
    events = [e for e in store.values() if is_upcoming(e)]
    def _sort_key(e: dict) -> date:
        parsed = _parse_event_date(e.get("date", ""))
        return parsed or date(9999, 12, 31)
    return sorted(events, key=_sort_key)

# ── Brave Search API ──────────────────────────────────────────────────────────

def search_brave_api(query: str, count: int = 10) -> list[dict]:
    """
    Call Brave's Web Search API — searches the entire web, no CAPTCHA.
    Free tier: 2000 queries/month.  Max 20 results per call.
    Silently returns [] if BRAVE_API_KEY is not set.
    """
    if not BRAVE_API_KEY:
        return []

    print(f"  [Brave API] {query}")
    url = "https://api.search.brave.com/res/v1/web/search"
    headers = {
        "Accept":             "application/json",
        "Accept-Encoding":    "gzip",
        "X-Subscription-Token": BRAVE_API_KEY,
    }
    params = {
        "q":          query,
        "count":      min(count, 20),
        "freshness":  "py",   # past year — keeps results current
        "text_decorations": 0,
    }

    try:
        resp = requests.get(url, headers=headers, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
    except Exception as exc:
        print(f"    [Brave API] error: {exc}")
        return []

    results = []
    for item in data.get("web", {}).get("results", []):
        snippet = item.get("description", "")
        results.append({
            "title":    item.get("title", "").strip(),
            "date":     extract_date(snippet),
            "location": extract_location(snippet),
            "link":     item.get("url", ""),
            "source":   "Brave Search API",
            "query":    query,
            "snippet":  snippet,
        })
    return results

# ── Eventbrite Scraper ────────────────────────────────────────────────────────

async def scrape_eventbrite(page, keyword: str) -> list[dict]:
    results = []
    print(f"  [Eventbrite] {keyword}")

    # Use the conference/summit category filter in the URL
    url = (
        f"https://www.eventbrite.com/d/united-states/{keyword.replace(' ', '-')}/"
        f"?subcategory=6003"  # 6003 = Conferences
    )
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        return results

    await page.wait_for_timeout(3000)

    # Scroll to trigger lazy-loaded cards
    for _ in range(3):
        await page.keyboard.press("End")
        await page.wait_for_timeout(1500)

    # Primary selectors (Eventbrite updates its DOM regularly — fallback chain)
    selectors = [
        "div[data-testid='event-card']",
        "article.eds-event-card-content",
        "section.search-results-panel-content a[href*='eventbrite.com/e/']",
    ]

    for sel in selectors:
        cards = page.locator(sel)
        if await cards.count() > 0:
            break

    card_count = await cards.count()

    # Final fallback: any Eventbrite event link on the page
    if card_count == 0:
        links = page.locator("a[href*='eventbrite.com/e/']")
        for i in range(min(await links.count(), 25)):
            lnk = links.nth(i)
            try:
                href = await lnk.get_attribute("href") or ""
                text = (await lnk.inner_text()).strip()
                lines = [l.strip() for l in text.splitlines() if l.strip()]
                if not lines:
                    continue
                title = lines[0]
                full_text = " ".join(lines)
                results.append({
                    "title":    title,
                    "date":     extract_date(full_text),
                    "location": extract_location(full_text),
                    "link":     href.split("?")[0],
                    "source":   "Eventbrite",
                    "query":    keyword,
                    "snippet":  full_text,
                })
            except Exception:
                continue
        return results

    for i in range(min(card_count, 25)):
        card = cards.nth(i)
        try:
            title_el = card.locator("h2, h3, .eds-event-card__formatted-name").first
            title = (await title_el.inner_text()).strip() if await title_el.count() > 0 else ""

            link_el = card.locator("a[href*='eventbrite.com/e/']").first
            href = (await link_el.get_attribute("href") or "") if await link_el.count() > 0 else ""

            date_el = card.locator(
                "p[data-testid='event-card-date'], .eds-event-card-content__sub-title, time"
            ).first
            date_text = (await date_el.inner_text()).strip() if await date_el.count() > 0 else ""

            loc_el = card.locator(
                "p[data-testid='event-card-location'], .card-text--truncated"
            ).first
            location = (await loc_el.inner_text()).strip() if await loc_el.count() > 0 else ""

            full_text = (await card.inner_text()).strip()

            if title:
                results.append({
                    "title":    title,
                    "date":     date_text or extract_date(full_text),
                    "location": location or extract_location(full_text),
                    "link":     href.split("?")[0] if href else "",
                    "source":   "Eventbrite",
                    "query":    keyword,
                    "snippet":  full_text[:300],
                })
        except Exception:
            continue

    return results

# ── Meetup Scraper ────────────────────────────────────────────────────────────

async def scrape_meetup(page, keyword: str) -> list[dict]:
    results = []
    print(f"  [Meetup] {keyword}")
    url = (
        f"https://www.meetup.com/find/?keywords={keyword.replace(' ', '%20')}"
        f"&source=EVENTS&eventType=inPerson"
    )
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        return results

    await page.wait_for_timeout(3000)

    for _ in range(3):
        await page.keyboard.press("End")
        await page.wait_for_timeout(1500)

    cards = page.locator(
        "div[data-testid='categoryResults-eventCard'], article[data-element-name='event-card']"
    )
    card_count = await cards.count()

    if card_count == 0:
        links = page.locator("a[href*='/events/']")
        for i in range(min(await links.count(), 20)):
            lnk = links.nth(i)
            try:
                href = await lnk.get_attribute("href") or ""
                text = (await lnk.inner_text()).strip()
                if len(text) < 5:
                    continue
                full_href = (
                    f"https://www.meetup.com{href}" if href.startswith("/") else href
                )
                results.append({
                    "title":    text.splitlines()[0].strip(),
                    "date":     extract_date(text),
                    "location": extract_location(text),
                    "link":     full_href,
                    "source":   "Meetup",
                    "query":    keyword,
                    "snippet":  text[:300],
                })
            except Exception:
                continue
        return results

    for i in range(min(card_count, 20)):
        card = cards.nth(i)
        try:
            title_el = card.locator("h3, h2, [data-testid='event-card-title']").first
            title = (await title_el.inner_text()).strip() if await title_el.count() > 0 else ""

            link_el = card.locator("a[href*='/events/']").first
            href = (await link_el.get_attribute("href") or "") if await link_el.count() > 0 else ""
            if href.startswith("/"):
                href = f"https://www.meetup.com{href}"

            date_el = card.locator("time, [data-testid='event-card-date']").first
            date_text = ""
            if await date_el.count() > 0:
                date_text = (
                    await date_el.get_attribute("datetime")
                    or await date_el.inner_text()
                    or ""
                ).strip()

            loc_el = card.locator("[data-testid='event-card-venue'], .venueDisplay").first
            location = (await loc_el.inner_text()).strip() if await loc_el.count() > 0 else ""

            full_text = (await card.inner_text()).strip()

            if title:
                results.append({
                    "title":    title,
                    "date":     date_text or extract_date(full_text),
                    "location": location or extract_location(full_text),
                    "link":     href,
                    "source":   "Meetup",
                    "query":    keyword,
                    "snippet":  full_text[:300],
                })
        except Exception:
            continue

    return results

# ── Luma Scraper ──────────────────────────────────────────────────────────────

async def scrape_luma(page, keyword: str) -> list[dict]:
    """Scrape lu.ma search results — a popular platform for tech community events."""
    results = []
    print(f"  [Luma] {keyword}")
    url = f"https://lu.ma/search?q={keyword.replace(' ', '+')}"
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        return results

    await page.wait_for_timeout(3000)

    for _ in range(2):
        await page.keyboard.press("End")
        await page.wait_for_timeout(1500)

    # Luma event cards
    cards = page.locator("a[href*='lu.ma/'], div[data-testid='event-row']")
    card_count = await cards.count()

    for i in range(min(card_count, 25)):
        card = cards.nth(i)
        try:
            href = await card.get_attribute("href") or ""
            text = (await card.inner_text()).strip()
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            if not lines:
                continue
            title = lines[0]
            full_text = " ".join(lines)
            if not href.startswith("http"):
                href = f"https://lu.ma{href}"
            results.append({
                "title":    title,
                "date":     extract_date(full_text),
                "location": extract_location(full_text),
                "link":     href.split("?")[0],
                "source":   "Luma",
                "query":    keyword,
                "snippet":  full_text[:300],
            })
        except Exception:
            continue

    return results

# ── Devpost Scraper ───────────────────────────────────────────────────────────

async def scrape_devpost(page, keyword: str) -> list[dict]:
    """Scrape Devpost hackathon listings — the primary platform for in-person hackathons."""
    results = []
    print(f"  [Devpost] {keyword}")
    url = (
        f"https://devpost.com/hackathons"
        f"?search={keyword.replace(' ', '+')}"
        f"&challenge_type[]=in-person"
        f"&order_by=recently-added"
    )
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        return results

    await page.wait_for_timeout(3000)

    # Primary card selector confirmed by scraper research
    cards = page.locator("article.challenge-listing")
    card_count = await cards.count()

    if card_count == 0:
        # Fallback: any individual hackathon subdomain links on the page
        links = page.locator("a[href*='.devpost.com']")
        for i in range(min(await links.count(), 20)):
            lnk = links.nth(i)
            try:
                href  = await lnk.get_attribute("href") or ""
                if not href or "devpost.com/hackathons" in href:
                    continue
                text  = (await lnk.inner_text()).strip()
                lines = [l.strip() for l in text.splitlines() if l.strip()]
                if not lines:
                    continue
                full_text = " ".join(lines)
                results.append({
                    "title":    lines[0],
                    "date":     extract_date(full_text),
                    "location": extract_location(full_text),
                    "link":     href.split("?")[0],
                    "source":   "Devpost",
                    "query":    keyword,
                    "snippet":  full_text[:300],
                })
            except Exception:
                continue
        return results

    for i in range(min(card_count, 20)):
        card = cards.nth(i)
        try:
            title_el = card.locator("h2, h3, .title, [class*='title']").first
            title    = (await title_el.inner_text()).strip() if await title_el.count() > 0 else ""

            link_el  = card.locator("a[href*='.devpost.com']").first
            href     = (await link_el.get_attribute("href") or "") if await link_el.count() > 0 else ""

            date_el  = card.locator("span.value.date-range, time, [class*='date']").first
            date_text = (await date_el.inner_text()).strip() if await date_el.count() > 0 else ""

            full_text = (await card.inner_text()).strip()

            if title:
                results.append({
                    "title":    title,
                    "date":     date_text or extract_date(full_text),
                    "location": extract_location(full_text),
                    "link":     href.split("?")[0] if href else "",
                    "source":   "Devpost",
                    "query":    keyword,
                    "snippet":  full_text[:300],
                })
        except Exception:
            continue

    return results


# ── Confs.tech Fetcher ────────────────────────────────────────────────────────

def fetch_confs_tech(year: int) -> list[dict]:
    """
    Pull conference data from the Confs.tech open dataset (GitHub-hosted JSON).
    No scraping — clean structured data, filtered to North America.
    """
    results = []
    base = (
        "https://raw.githubusercontent.com/tech-conferences/conference-data"
        f"/main/conferences/{year}"
    )
    for topic in CONFS_TECH_TOPICS:
        try:
            resp = requests.get(f"{base}/{topic}.json", timeout=10)
            if resp.status_code != 200:
                continue
            for conf in resp.json():
                if conf.get("online"):
                    continue
                country = conf.get("country", "")
                if country and country not in _CONFS_TECH_NA_COUNTRIES:
                    continue
                city  = conf.get("city", "")
                state = conf.get("stateProvince", "") or country
                loc   = f"{city}, {state}".strip(", ") if city else state
                results.append({
                    "title":    conf.get("name", ""),
                    "date":     conf.get("startDate", ""),
                    "location": loc,
                    "link":     conf.get("url", ""),
                    "source":   "Confs.tech",
                    "query":    topic,
                    "snippet":  conf.get("name", ""),
                })
        except Exception as exc:
            print(f"    [Confs.tech] {topic}/{year}: {exc}")
    return results


# ── Excel Export ──────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
_LINK_FONT    = Font(color="1155CC", underline="single")
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP         = Alignment(vertical="top", wrap_text=True)


def export_to_excel(results: list[dict], filename: str) -> None:
    wb = Workbook()

    # ── Events sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Events"

    headers    = ["Title", "Date", "Location", "Community", "Link", "Source", "Search Query"]
    col_widths = [50,      22,     30,          20,           55,     16,        35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Community colour fills for quick visual scanning
    _COMM_FILLS = {
        "Black Tech":     PatternFill("solid", fgColor="C8A2C8"),   # lilac
        "Women in Tech":  PatternFill("solid", fgColor="FADADD"),   # rose
        "LGBTQ+ Tech":    PatternFill("solid", fgColor="B5EAD7"),   # mint
        "General":        None,
    }

    for row_idx, event in enumerate(results, start=2):
        community = event.get("community", "General")
        link      = event.get("link", "")
        values = [
            event.get("title", ""),
            event.get("date", ""),
            event.get("location", ""),
            community,
            link,
            event.get("source", ""),
            event.get("query", ""),
        ]
        # Pick fill: community colour on even rows, alt grey on odd (no community colour)
        row_fill = _COMM_FILLS.get(community.split(" / ")[0]) or (
            _ALT_FILL if row_idx % 2 == 0 else None
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _WRAP
            if row_fill:
                cell.fill = row_fill
            # Make links clickable (link is now column 5)
            if col_idx == 5 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font      = _LINK_FONT

    # ── Summary sheet ─────────────────────────────────────────────────────────
    sw = wb.create_sheet("Summary")
    sw.column_dimensions["A"].width = 30
    sw.column_dimensions["B"].width = 12

    title_cell = sw.cell(row=1, column=1, value="Tech Events — Scrape Summary")
    title_cell.font = Font(bold=True, size=14)

    sw.cell(row=3, column=1, value="Scraped On")
    sw.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    sw.cell(row=4, column=1, value="Total Results")
    sw.cell(row=4, column=2, value=len(results))

    # By source
    sw.cell(row=6, column=1, value="Source").font = Font(bold=True)
    sw.cell(row=6, column=2, value="Count").font  = Font(bold=True)
    source_counts: dict[str, int] = {}
    for r in results:
        src = r.get("source", "Unknown")
        source_counts[src] = source_counts.get(src, 0) + 1
    row = 7
    for src, cnt in source_counts.items():
        sw.cell(row=row, column=1, value=src)
        sw.cell(row=row, column=2, value=cnt)
        row += 1

    # By community
    row += 1
    sw.cell(row=row, column=1, value="Community").font = Font(bold=True)
    sw.cell(row=row, column=2, value="Count").font     = Font(bold=True)
    row += 1
    comm_counts: dict[str, int] = {}
    for r in results:
        for tag in r.get("community", "General").split(" / "):
            comm_counts[tag] = comm_counts.get(tag, 0) + 1
    for comm, cnt in sorted(comm_counts.items()):
        sw.cell(row=row, column=1, value=comm)
        sw.cell(row=row, column=2, value=cnt)
        row += 1

    wb.save(filename)
    print(f"Excel saved → {filename}")

# ── Google Sheets Export (optional) ──────────────────────────────────────────

def export_to_sheets(results: list[dict], sheet_name: str) -> str:
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        from google.oauth2 import service_account
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
    except ImportError:
        print("gspread not installed — skipping Sheets export.")
        return ""

    print(f"\nCreating Google Sheet: {sheet_name}")

    if SHEETS_AUTH_METHOD == "service_account":
        if not Path(SERVICE_ACCOUNT_FILE).exists():
            raise FileNotFoundError(f"Missing {SERVICE_ACCOUNT_FILE}")
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SHEET_SCOPES
        )
    else:
        creds = None
        if Path(TOKEN_FILE).exists():
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SHEET_SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                if not Path(CREDENTIALS_FILE).exists():
                    raise FileNotFoundError(f"Missing {CREDENTIALS_FILE}")
                flow = InstalledAppFlow.from_client_secrets_file(
                    CREDENTIALS_FILE, SHEET_SCOPES
                )
                creds = flow.run_local_server(port=0)
            with open(TOKEN_FILE, "w") as f:
                f.write(creds.to_json())

    gc           = gspread.authorize(creds)
    spreadsheet  = gc.create(sheet_name)
    ws           = spreadsheet.sheet1
    ws.update_title("All Events")

    headers = ["Title", "Date", "Location", "Community", "Link", "Source", "Search Query"]
    rows    = [headers] + [
        [
            e.get("title", ""), e.get("date", ""), e.get("location", ""),
            e.get("community", "General"), e.get("link", ""),
            e.get("source", ""), e.get("query", ""),
        ]
        for e in results
    ]
    ws.update(rows, "A1")

    ws.format("A1:G1", {
        "backgroundColor": {"red": 0.102, "green": 0.102, "blue": 0.180},
        "textFormat": {
            "bold": True,
            "foregroundColor": {"red": 1, "green": 1, "blue": 1},
            "fontSize": 11,
        },
        "horizontalAlignment": "CENTER",
    })

    spreadsheet.batch_update({"requests": [
        {"updateSheetProperties": {
            "properties": {"sheetId": ws.id, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount",
        }},
        {"autoResizeDimensions": {
            "dimensions": {"sheetId": ws.id, "dimension": "COLUMNS",
                           "startIndex": 0, "endIndex": 7},
        }},
    ]})

    # Summary tab
    sw = spreadsheet.add_worksheet(title="Summary", rows=20, cols=4)
    source_counts: dict[str, int] = {}
    for r in results:
        src = r.get("source", "Unknown")
        source_counts[src] = source_counts.get(src, 0) + 1

    summary_rows = [
        ["Black Tech Events — Scrape Summary", "", "", ""],
        [""],
        ["Scraped On", datetime.now().strftime("%Y-%m-%d %H:%M"), "", ""],
        ["Total Results", str(len(results)), "", ""],
        [""],
        ["Source", "Count", "", ""],
    ] + [[src, str(cnt), "", ""] for src, cnt in source_counts.items()]

    sw.update(summary_rows, "A1")
    sw.format("A1",  {"textFormat": {"bold": True, "fontSize": 14}})
    sw.format("A6:B6", {"textFormat": {"bold": True}})

    try:
        spreadsheet.share(None, perm_type="anyone", role="reader")
    except Exception:
        pass

    url = spreadsheet.url
    print(f"Google Sheet → {url}")
    return url

# ── HTML Export (GitHub Pages) ───────────────────────────────────────────────

def export_to_html(results: list[dict], filename: str) -> None:
    """
    Generate a self-contained HTML page for GitHub Pages.
    Includes live search and community filter buttons — no external dependencies.
    """
    Path(filename).parent.mkdir(parents=True, exist_ok=True)

    events_data = [
        {
            "title":      e.get("title", ""),
            "date":       e.get("date", ""),
            "location":   e.get("location", ""),
            "community":  e.get("community", "General"),
            "link":       e.get("link", ""),
            "source":     e.get("source", ""),
            "first_seen": e.get("first_seen", ""),
        }
        for e in results
    ]
    # Escape </script> so it can't break out of the <script> block
    events_json = json.dumps(events_data, ensure_ascii=False).replace(
        "</script>", r"<\/script>"
    )

    updated  = datetime.now().strftime("%B %d, %Y at %H:%M UTC")
    count    = len(results)
    repo     = os.getenv("GITHUB_REPOSITORY", "")
    gh_link  = (
        f'&middot; <a href="https://github.com/{repo}" target="_blank" rel="noopener">'
        f'View on GitHub</a>'
    ) if repo else ""

    page = f"""\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Black Tech Events</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      background: #f4f5f7;
      color: #1a1a2e;
    }}

    /* ── Header ── */
    header {{
      background: #1a1a2e;
      color: #fff;
      padding: 2.5rem 1.5rem 2rem;
      text-align: center;
    }}
    header h1   {{ font-size: 2rem; font-weight: 700; letter-spacing: -0.01em; margin-bottom: 0.3rem; }}
    header .sub {{ opacity: 0.7; font-size: 0.95rem; margin-bottom: 0.25rem; }}
    header .meta {{ font-size: 0.8rem; opacity: 0.5; }}

    /* ── Sticky controls bar ── */
    .controls {{
      position: sticky; top: 0; z-index: 50;
      background: #fff;
      border-bottom: 1px solid #e2e5ea;
      padding: 0.65rem 1.5rem;
      display: flex; flex-wrap: wrap; gap: 0.5rem; align-items: center;
      box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }}
    #search {{
      flex: 1; min-width: 160px;
      padding: 0.38rem 0.75rem;
      border: 1px solid #ced4da; border-radius: 6px;
      font-size: 0.88rem; outline: none; color: #1a1a2e;
    }}
    #search:focus {{ border-color: #1a1a2e; }}
    .filters {{ display: flex; flex-wrap: wrap; gap: 0.3rem; }}
    .f {{
      padding: 0.28rem 0.7rem;
      border: 1px solid #ced4da; border-radius: 20px;
      background: #fff; cursor: pointer;
      font-size: 0.8rem; font-weight: 500; color: #555;
      transition: background 0.12s, color 0.12s, border-color 0.12s;
    }}
    .f:hover {{ background: #f0f0f5; }}
    .f.on {{ background: #1a1a2e; color: #fff; border-color: #1a1a2e; }}
    .f[data-v="Black Tech"].on    {{ background: #6b4c8a; border-color: #6b4c8a; }}
    .f[data-v="Women in Tech"].on {{ background: #b05070; border-color: #b05070; }}
    .f[data-v="LGBTQ+ Tech"].on  {{ background: #2e8f68; border-color: #2e8f68; }}
    #count {{ margin-left: auto; font-size: 0.8rem; color: #888; white-space: nowrap; }}

    /* ── Table ── */
    .wrap {{ padding: 1.25rem 1.5rem 3rem; }}
    table {{
      width: 100%; border-collapse: collapse;
      background: #fff; border-radius: 8px;
      overflow: hidden;
      box-shadow: 0 1px 4px rgba(0,0,0,0.07);
    }}
    th {{
      background: #1a1a2e; color: #fff;
      font-size: 0.78rem; font-weight: 600;
      text-transform: uppercase; letter-spacing: 0.06em;
      padding: 0.6rem 0.9rem; text-align: left;
    }}
    td {{
      padding: 0.55rem 0.9rem;
      font-size: 0.875rem;
      border-bottom: 1px solid #f0f0f0;
      vertical-align: top;
    }}
    tr:last-child td {{ border-bottom: none; }}
    tr.r-black td {{ background: #f8f4f8; }}
    tr.r-women td {{ background: #fdf6f7; }}
    tr.r-lgbtq td {{ background: #f2faf7; }}
    tr:hover td {{ filter: brightness(0.97); }}

    /* ── Community tags ── */
    .tag {{
      display: inline-block; padding: 0.18rem 0.5rem;
      border-radius: 10px; font-size: 0.72rem; font-weight: 600;
      white-space: nowrap; margin: 0.1rem 0;
    }}
    .t-black   {{ background: #c8a2c8; color: #3d1a5e; }}
    .t-women   {{ background: #fadadd; color: #7a1a3a; }}
    .t-lgbtq   {{ background: #b5ead7; color: #0f5c3a; }}
    .t-general {{ background: #e9ecef; color: #495057; }}

    a {{ color: #1a1a2e; text-decoration: none; font-weight: 500; }}
    a:hover {{ color: #4a4aff; text-decoration: underline; }}

    .empty {{
      text-align: center; padding: 3rem 1rem;
      color: #888; font-size: 0.9rem;
    }}

    .new-badge {{
      display: inline-block; padding: 0.1rem 0.42rem;
      border-radius: 4px; font-size: 0.64rem; font-weight: 700;
      background: #22c55e; color: #fff; vertical-align: middle;
      margin-left: 0.35rem; letter-spacing: 0.04em;
    }}

    footer {{
      text-align: center; padding: 1.5rem;
      font-size: 0.78rem; color: #aaa;
      border-top: 1px solid #e2e5ea;
    }}
    footer a {{ color: #888; }}

    /* ── Responsive — hide location column on small screens ── */
    @media (max-width: 600px) {{
      .c-loc {{ display: none; }}
      .wrap {{ padding: 0.75rem 0.5rem 2rem; }}
      .controls {{ padding: 0.5rem 0.75rem; }}
    }}
  </style>
</head>
<body>

<header>
  <h1>Black Tech Events</h1>
  <p class="sub">Upcoming tech conferences, summits &amp; hackathons</p>
  <p class="meta">Last updated: {updated} &middot; {count} events</p>
</header>

<div class="controls">
  <input id="search" type="text" placeholder="Search events, locations&hellip;" />
  <div class="filters">
    <button class="f on" data-v="all">All</button>
    <button class="f" data-v="Black Tech">Black Tech</button>
    <button class="f" data-v="Women in Tech">Women in Tech</button>
    <button class="f" data-v="LGBTQ+ Tech">LGBTQ+ Tech</button>
    <button class="f" data-v="General">General</button>
  </div>
  <span id="count"></span>
</div>

<div class="wrap">
  <table id="tbl">
    <thead>
      <tr>
        <th style="min-width:220px">Event</th>
        <th style="min-width:110px">Date</th>
        <th class="c-loc" style="min-width:150px">Location</th>
        <th style="min-width:140px">Community</th>
      </tr>
    </thead>
    <tbody id="tbody"></tbody>
  </table>
  <p class="empty" id="empty" style="display:none">
    No events match — try a different filter or search term.
  </p>
</div>

<footer>
  Collected weekly via Brave Search API, Eventbrite, Meetup &amp; Luma
  {gh_link}
</footer>

<script>
const DATA = {events_json};

const TAG_CLASS = {{
  'Black Tech':    't-black',
  'Women in Tech': 't-women',
  'LGBTQ+ Tech':   't-lgbtq',
  'General':       't-general',
}};

function renderTags(community) {{
  return community.split(' / ')
    .map(t => `<span class="tag ${{TAG_CLASS[t] || 't-general'}}">${{esc(t)}}</span>`)
    .join(' ');
}}

function rowClass(community) {{
  if (community.includes('Black Tech'))    return 'r-black';
  if (community.includes('Women in Tech')) return 'r-women';
  if (community.includes('LGBTQ+'))        return 'r-lgbtq';
  return '';
}}

function esc(s) {{
  return String(s)
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;');
}}

const TODAY_ISO = new Date().toISOString().slice(0, 10);
function isNew(firstSeen) {{
  if (!firstSeen) return false;
  return (new Date(TODAY_ISO) - new Date(firstSeen)) / 86400000 <= {NEW_BADGE_DAYS};
}}

let activeFilter = 'all';
let searchText   = '';

function render() {{
  const filtered = DATA.filter(e => {{
    const matchFilter = activeFilter === 'all' || e.community.includes(activeFilter);
    const q = searchText.toLowerCase();
    const matchSearch = !q ||
      e.title.toLowerCase().includes(q) ||
      (e.location || '').toLowerCase().includes(q) ||
      (e.date || '').toLowerCase().includes(q) ||
      e.community.toLowerCase().includes(q);
    return matchFilter && matchSearch;
  }});

  document.getElementById('tbody').innerHTML = filtered.map(e => `
    <tr class="${{rowClass(e.community)}}">
      <td>${{e.link
        ? `<a href="${{esc(e.link)}}" target="_blank" rel="noopener noreferrer">${{esc(e.title)}}</a>`
        : esc(e.title)}}${{isNew(e.first_seen) ? ' <span class="new-badge">NEW</span>' : ''}}</td>
      <td>${{esc(e.date || '\u2014')}}</td>
      <td class="c-loc">${{esc(e.location || '\u2014')}}</td>
      <td>${{renderTags(e.community)}}</td>
    </tr>`).join('');

  document.getElementById('count').textContent =
    filtered.length + ' of ' + DATA.length + ' events';
  document.getElementById('empty').style.display = filtered.length ? 'none' : '';
  document.getElementById('tbl').style.display   = filtered.length ? ''     : 'none';
}}

document.querySelectorAll('.f').forEach(btn => {{
  btn.addEventListener('click', () => {{
    document.querySelectorAll('.f').forEach(b => b.classList.remove('on'));
    btn.classList.add('on');
    activeFilter = btn.dataset.v;
    render();
  }});
}});

document.getElementById('search').addEventListener('input', e => {{
  searchText = e.target.value;
  render();
}});

render();
</script>
</body>
</html>"""

    Path(filename).write_text(page, encoding="utf-8")
    print(f"HTML  saved → {filename}")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main() -> None:
    # ── 0. Load persistent event store ───────────────────────────────────────
    today = _TODAY.isoformat()
    store = load_event_store()
    store_count_before = len(store)

    all_results: list[dict] = []

    # ── 1. Brave Search API (no browser needed, whole-web coverage) ──────────
    if BRAVE_API_KEY:
        # Tier 1 — site-targeted queries (highest precision: direct event pages)
        print("\n── Brave Search API — site-targeted ───────────────")
        site_hits: list[dict] = []
        for query in BRAVE_SITE_QUERIES:
            hits = search_brave_api(query)
            for h in hits:
                h["_tier"] = "site"
            site_hits.extend(hits)
        all_results.extend(site_hits)
        print(f"  → {len(site_hits)} site-targeted results")

        # Tier 2 — general whole-web queries (broader reach, stricter filter later)
        print("\n── Brave Search API — whole-web ───────────────────")
        web_hits: list[dict] = []
        for query in BRAVE_QUERIES:
            hits = search_brave_api(query)
            for h in hits:
                h["_tier"] = "web"
            web_hits.extend(hits)
        all_results.extend(web_hits)
        print(f"  → {len(web_hits)} whole-web results")
    else:
        print(
            "\n[SKIP] Brave API key not set — add BRAVE_API_KEY to .env "
            "to enable whole-web search results."
        )

    # ── 2. Eventbrite + Meetup + Luma (Playwright) ────────────────────────────
    print("\n── Browser-based scrapers ─────────────────────────")
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )
        page = await context.new_page()

        print("\n── Eventbrite ─────────────────────────────────────")
        for keyword in EVENTBRITE_QUERIES:
            hits = await scrape_eventbrite(page, keyword)
            all_results.extend(hits)
            await page.wait_for_timeout(2000)

        print("\n── Meetup ──────────────────────────────────────────")
        for keyword in MEETUP_QUERIES:
            hits = await scrape_meetup(page, keyword)
            all_results.extend(hits)
            await page.wait_for_timeout(2000)

        print("\n── Luma (lu.ma) ────────────────────────────────────")
        for keyword in LUMA_QUERIES:
            hits = await scrape_luma(page, keyword)
            all_results.extend(hits)
            await page.wait_for_timeout(2000)

        print("\n── Devpost ─────────────────────────────────────────")
        for keyword in DEVPOST_QUERIES:
            hits = await scrape_devpost(page, keyword)
            all_results.extend(hits)
            await page.wait_for_timeout(2000)

        await browser.close()

    # ── 2b. Confs.tech (no browser — plain HTTP JSON fetch) ───────────────────
    print("\n── Confs.tech ──────────────────────────────────────")
    for year in (_YEAR, _YEAR + 1):
        hits = fetch_confs_tech(year)
        all_results.extend(hits)
        print(f"  → {len(hits)} results for {year}")

    # ── 3. Deduplicate ────────────────────────────────────────────────────────
    deduped = deduplicate(all_results)

    # ── 4. Relevance filter ───────────────────────────────────────────────────
    # Scraper results trust the source; general Brave web results get strict mode.
    relevant = [
        e for e in deduped
        if is_relevant(e, strict=(e.get("_tier") == "web"))
    ]
    relevant = relevant if relevant else deduped  # safety fallback

    # ── 5. Filter to North America ────────────────────────────────────────────
    na_only = [e for e in relevant if _is_north_america(e)]
    intl_ct = len(relevant) - len(na_only)

    # ── 6. Drop past events ───────────────────────────────────────────────────
    upcoming = [e for e in na_only if is_upcoming(e)]
    past_ct  = len(na_only) - len(upcoming)

    # ── 7. Tag community ──────────────────────────────────────────────────────
    for e in upcoming:
        e["community"] = classify_community(e)

    final = upcoming

    print(f"\nRaw results        : {len(all_results)}")
    print(f"After dedup        : {len(deduped)}")
    print(f"After relevance    : {len(relevant)}")
    print(f"Non-NA dropped     : {intl_ct}")
    print(f"Past events dropped: {past_ct}")
    print(f"Found this run     : {len(final)} upcoming events")
    if final:
        comm_summary: dict[str, int] = {}
        for e in final:
            for tag in e["community"].split(" / "):
                comm_summary[tag] = comm_summary.get(tag, 0) + 1
        for tag, cnt in sorted(comm_summary.items()):
            print(f"  {tag:<20}: {cnt}")
    else:
        print("\nWARNING: No events found this run.", file=sys.stderr)

    # ── 8. Merge into persistent store, expire stale entries, save ───────────
    store = merge_into_store(store, final, today)
    store = expire_past_events(store, today)
    save_event_store(store)
    print(f"Event store        : {store_count_before} → {len(store)} events")

    if not store:
        print("ERROR: Event store is empty — nothing to export.", file=sys.stderr)
        raise SystemExit(1)

    # ── 9. Export from full store (includes events from previous runs) ────────
    export_list  = store_to_list(store)
    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file   = f"black_tech_events_{timestamp}.xlsx"
    export_to_excel(export_list, excel_file)
    export_to_html(export_list, "docs/index.html")

    if ENABLE_SHEETS:
        sheet_name = f"Black Tech Events {datetime.now().strftime('%Y-%m-%d')}"
        try:
            export_to_sheets(export_list, sheet_name)
        except Exception as exc:
            print(f"Sheets export failed: {exc}")

    print(f"\nDone! {len(export_list)} events in store, exported to {excel_file}")


if __name__ == "__main__":
    asyncio.run(main())
